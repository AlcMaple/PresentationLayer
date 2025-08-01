from typing import List, Optional, Dict, Any, Tuple
from sqlmodel import Session, select, and_
from sqlalchemy import func
from datetime import datetime, timezone
from openpyxl import Workbook
from io import BytesIO
from openpyxl.styles import Font, PatternFill
import traceback


from models.paths import Paths
from schemas.paths import PathsCreate, PathsUpdate, PathConditions, PathsResponse
from services.base_crud import BaseCRUDService, PageParams
from models import (
    Categories,
    AssessmentUnit,
    BridgeTypes,
    BridgeParts,
    BridgeStructures,
    BridgeComponentTypes,
    BridgeComponentForms,
    BridgeDiseases,
    BridgeScales,
    BridgeQualities,
    BridgeQuantities,
)
from exceptions import DuplicateException, NotFoundException, ValidationException
from utils import (
    create_reference_data_sheet,
    create_help_sheet,
    validate_excel_data,
    get_reference_data,
    match_name_to_code,
)


# 需要自定义类，不能用工厂函数，只能继承
class PathsService(BaseCRUDService[Paths, PathsCreate, PathsUpdate]):
    """路径服务类"""

    def __init__(self, session: Session):
        super().__init__(Paths, session)

    def get_paths_with_pagination(
        self, page_params: PageParams, conditions: Optional[PathConditions] = None
    ) -> Tuple[List[PathsResponse], int]:
        """
        分页查询路径数据

        Args:
            page_params: 分页参数
            conditions: 查询条件
        """
        try:
            # 构建基础查询
            statement = select(Paths)
            count_statement = select(func.count(Paths.id))

            # 构建过滤条件
            filter_conditions = []
            if hasattr(Paths, "is_active"):
                filter_conditions.append(Paths.is_active == True)

            if conditions:
                filter_conditions.extend(self._build_path_filter_conditions(conditions))

            # 应用过滤条件
            if filter_conditions:
                statement = statement.where(and_(*filter_conditions))
                count_statement = count_statement.where(and_(*filter_conditions))

            # 排序
            statement = statement.order_by(Paths.id)

            # 分页
            statement = statement.offset(page_params.offset).limit(page_params.size)

            # 执行查询
            results = self.session.exec(statement).all()
            total = self.session.exec(count_statement).first() or 0

            # 转换为PathsResponse格式
            paths_list = []
            for result in results:
                path_data = {
                    "id": result.id,
                    "code": result.code,
                    "name": result.name,
                }

                # 添加各个ID对应的关联数据
                path_data.update(self._get_related_data(result))
                paths_list.append(PathsResponse(**path_data))

            return paths_list, total

        except Exception as e:
            print(f"分页查询路径数据时出错: {e}")
            return [], 0

    def _get_related_data(self, path_result) -> Dict[str, Any]:
        """
        获取path记录中各个ID对应的关联数据
        """
        related_data = {}

        # 定义字段映射关系
        field_mappings = [
            ("category_id", Categories, "category"),
            ("assessment_unit_id", AssessmentUnit, "assessment_unit"),
            ("bridge_type_id", BridgeTypes, "bridge_type"),
            ("part_id", BridgeParts, "part"),
            ("structure_id", BridgeStructures, "structure"),
            ("component_type_id", BridgeComponentTypes, "component_type"),
            ("component_form_id", BridgeComponentForms, "component_form"),
            ("disease_id", BridgeDiseases, "disease"),
            ("scale_id", BridgeScales, "scale"),
            ("quality_id", BridgeQualities, "quality"),
            ("quantity_id", BridgeQuantities, "quantity"),
        ]

        for field_name, model_class, prefix in field_mappings:
            field_id = getattr(path_result, field_name, None)

            # 返回各个字段对应的 id
            related_data[field_name] = field_id
            if field_id:
                try:
                    # 标度表处理
                    if model_class == BridgeScales:
                        stmt = select(
                            BridgeScales.code,
                            BridgeScales.name,
                            BridgeScales.scale_type,
                            BridgeScales.scale_value,
                            BridgeScales.min_value,
                            BridgeScales.max_value,
                            BridgeScales.unit,
                            BridgeScales.display_text,
                        ).where(BridgeScales.id == field_id)
                        result = self.session.exec(stmt).first()
                        if result:
                            related_data[f"{prefix}_code"] = result[0]
                            related_data[f"{prefix}_name"] = result[1]
                            related_data["scale_type"] = result[2]
                            related_data["scale_value"] = result[3]
                            related_data["min_value"] = result[4]
                            related_data["max_value"] = result[5]
                            related_data["unit"] = result[6]
                            related_data["display_text"] = result[7]
                        else:
                            related_data[f"{prefix}_code"] = None
                            related_data[f"{prefix}_name"] = None
                            related_data["scale_type"] = None
                            related_data["scale_value"] = None
                            related_data["min_value"] = None
                            related_data["max_value"] = None
                            related_data["unit"] = None
                            related_data["display_text"] = None
                    else:
                        # 其他表的处理
                        stmt = select(model_class.code, model_class.name).where(
                            model_class.id == field_id
                        )
                        result = self.session.exec(stmt).first()
                        if result:
                            related_data[f"{prefix}_code"] = result[0]
                            related_data[f"{prefix}_name"] = result[1]
                        else:
                            related_data[f"{prefix}_code"] = None
                            related_data[f"{prefix}_name"] = None
                except Exception as e:
                    print(
                        f"查询模型 {model_class.__name__} 中ID为 {field_id} 的记录时出错: {e}"
                    )
                    if model_class == BridgeScales:
                        related_data[f"{prefix}_code"] = None
                        related_data[f"{prefix}_name"] = None
                        related_data["scale_type"] = None
                        related_data["scale_value"] = None
                        related_data["min_value"] = None
                        related_data["max_value"] = None
                        related_data["unit"] = None
                        related_data["display_text"] = None
                    else:
                        related_data[f"{prefix}_code"] = None
                        related_data[f"{prefix}_name"] = None
            else:
                # 字段为空时的处理
                if model_class == BridgeScales:
                    related_data[f"{prefix}_code"] = None
                    related_data[f"{prefix}_name"] = None
                    related_data["scale_type"] = None
                    related_data["scale_value"] = None
                    related_data["min_value"] = None
                    related_data["max_value"] = None
                    related_data["unit"] = None
                    related_data["display_text"] = None
                else:
                    related_data[f"{prefix}_code"] = None
                    related_data[f"{prefix}_name"] = None

        return related_data

    def get_filter_options(self) -> Dict[str, List[Dict[str, Any]]]:
        """
        获取Paths 表中存在的所有相关表的关联数据

        Returns:
            包含各表选项的字典
        """
        try:
            options = {}

            # 定义字段映射
            table_configs = [
                ("categories", Categories, "category", Paths.category_id),
                (
                    "assessment_units",
                    AssessmentUnit,
                    "assessment_unit",
                    Paths.assessment_unit_id,
                ),
                ("bridge_types", BridgeTypes, "bridge_type", Paths.bridge_type_id),
                ("bridge_parts", BridgeParts, "part", Paths.part_id),
                (
                    "bridge_structures",
                    BridgeStructures,
                    "structure",
                    Paths.structure_id,
                ),
                (
                    "bridge_component_types",
                    BridgeComponentTypes,
                    "component_type",
                    Paths.component_type_id,
                ),
                (
                    "bridge_component_forms",
                    BridgeComponentForms,
                    "component_form",
                    Paths.component_form_id,
                ),
                ("bridge_diseases", BridgeDiseases, "disease", Paths.disease_id),
                ("bridge_scales", BridgeScales, "scale", Paths.scale_id),
                ("bridge_qualities", BridgeQualities, "quality", Paths.quality_id),
                ("bridge_quantities", BridgeQuantities, "quantity", Paths.quantity_id),
            ]

            for table_name, model_class, option_key, path_field in table_configs:
                try:
                    # 查询在paths表中存在的ID
                    existing_ids_stmt = (
                        select(path_field).where(path_field.is_not(None)).distinct()
                    )
                    existing_ids = self.session.exec(existing_ids_stmt).all()

                    if existing_ids:
                        # 标度表处理
                        if model_class == BridgeScales:
                            # 查询标度表的完整信息
                            stmt = (
                                select(
                                    BridgeScales.code,
                                    BridgeScales.scale_type,
                                    BridgeScales.scale_value,
                                    BridgeScales.min_value,
                                    BridgeScales.max_value,
                                    BridgeScales.unit,
                                    BridgeScales.display_text,
                                )
                                .where(
                                    and_(
                                        BridgeScales.id.in_(existing_ids),
                                        BridgeScales.is_active == True,
                                    )
                                )
                                .order_by(BridgeScales.code)
                            )

                            results = self.session.exec(stmt).all()
                            scale_options = []

                            for row in results:
                                (
                                    code,
                                    scale_type,
                                    scale_value,
                                    min_value,
                                    max_value,
                                    unit,
                                    display_text,
                                ) = row

                                # 根据标度类型构建不同的name
                                if scale_type == "NUMERIC":
                                    display_name = str(scale_value)
                                elif scale_type == "RANGE":
                                    display_name = f"{min_value}-{max_value}{unit}"
                                elif scale_type == "TEXT":
                                    display_name = display_text
                                else:
                                    display_name = None

                                scale_options.append(
                                    {"code": code, "name": display_name}
                                )

                            options[option_key] = scale_options
                        else:
                            # 其他表的处理
                            stmt = (
                                select(model_class.code, model_class.name)
                                .where(
                                    and_(
                                        model_class.id.in_(existing_ids),
                                        model_class.is_active == True,
                                    )
                                )
                                .order_by(model_class.code)
                            )

                            result = self.session.exec(stmt).all()
                            options[option_key] = [
                                {"code": row[0], "name": row[1]} for row in result
                            ]
                    else:
                        options[option_key] = []

                except Exception as e:
                    print(f"查询表 {table_name} 选项时出错: {e}")
                    options[option_key] = []

            return options

        except Exception as e:
            print(f"获取过滤选项时出错: {e}")
            return {}

    def _build_path_filter_conditions(self, conditions: PathConditions) -> List[Any]:
        """
        构建路径查询过滤条件

        Args:
            conditions: 查询条件对象
        """
        filter_conditions = []

        try:
            # 按code过滤
            code_mappings = [
                # 前端传入category_code，对应基础表模型Categories，对比 Paths 表的 category_id 字段
                ("category_code", Categories, Paths.category_id),
                ("assessment_unit_code", AssessmentUnit, Paths.assessment_unit_id),
                ("bridge_type_code", BridgeTypes, Paths.bridge_type_id),
                ("part_code", BridgeParts, Paths.part_id),
                ("structure_code", BridgeStructures, Paths.structure_id),
                ("component_type_code", BridgeComponentTypes, Paths.component_type_id),
                ("component_form_code", BridgeComponentForms, Paths.component_form_id),
                ("disease_code", BridgeDiseases, Paths.disease_id),
                ("scale_code", BridgeScales, Paths.scale_id),
                ("quality_code", BridgeQualities, Paths.quality_id),
                ("quantity_code", BridgeQuantities, Paths.quantity_id),
            ]

            for condition_field, model_class, path_field in code_mappings:
                # 获取前端传入的code值
                code_value = getattr(conditions, condition_field, None)
                if code_value:
                    try:
                        # 找到对应的ID
                        stmt = select(model_class.id).where(
                            model_class.code == code_value
                        )
                        result = self.session.exec(stmt).first()
                        if result:
                            filter_conditions.append(path_field == result)
                    except Exception as e:
                        print(
                            f"查询 {model_class.__name__} 中code为 {code_value} 的记录时出错: {e}"
                        )

        except Exception as e:
            print(f"构建路径过滤条件时出错: {e}")

        return filter_conditions

    def get_options(self) -> Dict[str, List[Dict[str, Any]]]:
        """
        获取所有相关表的关联数据

        Returns:
            包含各表选项的字典
        """
        all_options = {}

        # 定义基础表的映射关系
        table_configs = [
            ("categories", Categories, "category"),
            ("assessment_units", AssessmentUnit, "assessment_unit"),
            ("bridge_types", BridgeTypes, "bridge_type"),
            ("bridge_parts", BridgeParts, "part"),
            ("bridge_structures", BridgeStructures, "structure"),
            ("bridge_component_types", BridgeComponentTypes, "component_type"),
            ("bridge_component_forms", BridgeComponentForms, "component_form"),
            ("bridge_diseases", BridgeDiseases, "disease"),
            ("bridge_scales", BridgeScales, "scale"),
            ("bridge_qualities", BridgeQualities, "quality"),
            ("bridge_quantities", BridgeQuantities, "quantity"),
        ]

        for option_key, model_class, _ in table_configs:
            try:
                # 对标度表进行特殊处理
                if model_class == BridgeScales:
                    stmt = (
                        select(
                            BridgeScales.id,
                            BridgeScales.code,
                            BridgeScales.scale_type,
                            BridgeScales.scale_value,
                            BridgeScales.min_value,
                            BridgeScales.max_value,
                            BridgeScales.unit,
                            BridgeScales.display_text,
                        )
                        .where(BridgeScales.is_active == True)
                        .order_by(BridgeScales.code)
                    )
                    results = self.session.exec(stmt).all()
                    scale_options = []

                    for row in results:
                        (
                            id_val,
                            code,
                            scale_type,
                            scale_value,
                            min_value,
                            max_value,
                            unit,
                            display_text,
                        ) = row

                        # 根据标度类型构建不同的name
                        if scale_type == "NUMERIC":
                            display_name = str(scale_value)
                        elif scale_type == "RANGE":
                            display_name = f"{min_value}-{max_value}{unit}"
                        elif scale_type == "TEXT":
                            display_name = display_text
                        else:
                            display_name = None

                        scale_options.append(
                            {"id": id_val, "code": code, "name": display_name}
                        )

                    all_options[option_key] = scale_options
                else:
                    # 其他表的处理
                    stmt = (
                        select(model_class.id, model_class.code, model_class.name)
                        .where(model_class.is_active == True)
                        .order_by(model_class.code)
                    )
                    results = self.session.exec(stmt).all()
                    all_options[option_key] = [
                        {"id": row[0], "code": row[1], "name": row[2]}
                        for row in results
                    ]
            except Exception as e:
                print(f"查询表 {option_key} 选项时出错: {e}")
                all_options[option_key] = []

        return all_options

    def get_by_id_with_details(self, id: int) -> Optional[PathsResponse]:
        """根据 id 获取paths单条记录数据"""
        try:
            # 构建查询语句
            stmt = select(Paths).where(Paths.id == id)
            if hasattr(Paths, "is_active"):
                stmt = stmt.where(Paths.is_active == True)

            result = self.session.exec(stmt).first()

            if not result:
                return None

            # 构建基础数据
            path_data = {
                "id": result.id,
                "code": result.code,
                "name": result.name,
            }

            # 添加各个ID对应的关联数据
            path_data.update(self._get_related_data(result))

            return PathsResponse(**path_data)

        except Exception as e:
            print(f"获取paths单条记录数据时出错: {e}")
            return None

    def create(self, obj_in: PathsCreate) -> Optional[PathsResponse]:
        """
        创建paths记录

        Args:
            obj_in: 创建路径参数
        """
        try:
            # 检查 code 和 name
            final_code = self.code_generator.assign_or_generate_code(
                "paths", obj_in.code
            )

            if obj_in.name:
                stmt = select(Paths).where(Paths.name == obj_in.name)
                if hasattr(Paths, "is_active"):
                    stmt = stmt.where(Paths.is_active == True)
                existing_name = self.session.exec(stmt).first()
                if existing_name:
                    raise DuplicateException(
                        resource="Paths", field="name", value=obj_in.name
                    )

            # 通过各种 code 找到对应的 ID
            path_data = {
                "code": final_code,
                "name": obj_in.name,
            }

            # code 到 ID 的映射关系
            code_to_id_mappings = [
                ("category_code", Categories, "category_id"),
                ("assessment_unit_code", AssessmentUnit, "assessment_unit_id"),
                ("bridge_type_code", BridgeTypes, "bridge_type_id"),
                ("part_code", BridgeParts, "part_id"),
                ("structure_code", BridgeStructures, "structure_id"),
                ("component_type_code", BridgeComponentTypes, "component_type_id"),
                ("component_form_code", BridgeComponentForms, "component_form_id"),
                ("disease_code", BridgeDiseases, "disease_id"),
                ("scale_code", BridgeScales, "scale_id"),
                ("quality_code", BridgeQualities, "quality_id"),
                ("quantity_code", BridgeQuantities, "quantity_id"),
            ]

            for code_field, model_class, id_field in code_to_id_mappings:
                code_value = getattr(obj_in, code_field, None)
                if code_value:
                    # 查找对应的 ID
                    stmt = select(model_class.id).where(
                        and_(
                            model_class.code == code_value,
                            model_class.is_active == True,
                        )
                    )
                    result = self.session.exec(stmt).first()
                    if result:
                        path_data[id_field] = result
                    else:
                        raise ValidationException(
                            f"找不到 {code_field} 为 '{code_value}' 的记录"
                        )

            # 检查记录的唯一性
            path_uniqueness_fields = [
                "category_id",
                "assessment_unit_id",
                "bridge_type_id",
                "part_id",
                "structure_id",
                "component_type_id",
                "component_form_id",
                "disease_id",
                "scale_id",
                "quality_id",
                "quantity_id",
            ]

            # 构建唯一性检查条件
            uniqueness_conditions = []
            for field in path_uniqueness_fields:
                field_value = path_data.get(field)
                if field_value is not None:
                    uniqueness_conditions.append(getattr(Paths, field) == field_value)
                else:
                    uniqueness_conditions.append(getattr(Paths, field).is_(None))

            # 检查是否存在完全相同的路径记录
            if uniqueness_conditions:
                stmt = select(Paths).where(and_(*uniqueness_conditions))
                if hasattr(Paths, "is_active"):
                    stmt = stmt.where(Paths.is_active == True)
                existing_path = self.session.exec(stmt).first()
                if existing_path:
                    raise DuplicateException(
                        resource="Paths",
                        field="path_combination",
                        value="相同的路径组合已存在",
                    )

            # 创建记录
            path_model = Paths(**path_data)
            self.session.add(path_model)
            self.session.commit()
            self.session.refresh(path_model)

            # 构建返回数据
            response_data = {
                "id": path_model.id,
                "code": path_model.code,
                "name": path_model.name,
            }

            # 添加各个ID对应的关联数据
            response_data.update(self._get_related_data(path_model))

            return PathsResponse(**response_data)

        except Exception as e:
            self.session.rollback()
            print(f"创建paths记录时出错: {e}")
            raise

    def update(self, id: int, obj_in: PathsUpdate) -> Optional[PathsResponse]:
        """
        更新记录

        Args:
            id: 路径ID
            obj_in: 更新路径参数
        """
        try:
            # 查询现有记录
            db_obj = self.get_by_id(id)
            if not db_obj:
                raise NotFoundException(resource="Paths", identifier=str(id))

            obj_data = obj_in.model_dump(exclude_unset=True)

            # 处理编码
            if "code" in obj_data:
                code_value = obj_data["code"]
                if not code_value or not code_value.strip():
                    # 如果编码为空，移除该字段，保持原编码不变
                    obj_data.pop("code")
                else:
                    code_value = code_value.strip()
                    # 检查编码重复（排除当前记录）
                    statement = select(Paths).where(
                        and_(Paths.code == code_value, Paths.id != id)
                    )
                    if hasattr(Paths, "is_active"):
                        statement = statement.where(Paths.is_active == True)
                    existing = self.session.exec(statement).first()
                    if existing:
                        raise DuplicateException(
                            resource="Paths", field="code", value=code_value
                        )
                    obj_data["code"] = code_value

            # 检查名称重复（排除当前记录）
            if "name" in obj_data:
                statement = select(Paths).where(
                    and_(Paths.name == obj_data["name"], Paths.id != id)
                )
                if hasattr(Paths, "is_active"):
                    statement = statement.where(Paths.is_active == True)
                existing = self.session.exec(statement).first()
                if existing:
                    raise DuplicateException(
                        resource="Paths", field="name", value=obj_data["name"]
                    )

            # code 到 id 的映射
            code_to_id_mappings = [
                ("category_code", Categories, "category_id"),
                ("assessment_unit_code", AssessmentUnit, "assessment_unit_id"),
                ("bridge_type_code", BridgeTypes, "bridge_type_id"),
                ("part_code", BridgeParts, "part_id"),
                ("structure_code", BridgeStructures, "structure_id"),
                ("component_type_code", BridgeComponentTypes, "component_type_id"),
                ("component_form_code", BridgeComponentForms, "component_form_id"),
                ("disease_code", BridgeDiseases, "disease_id"),
                ("scale_code", BridgeScales, "scale_id"),
                ("quality_code", BridgeQualities, "quality_id"),
                ("quantity_code", BridgeQuantities, "quantity_id"),
            ]

            for code_field, model_class, id_field in code_to_id_mappings:
                if code_field in obj_data:
                    code_value = obj_data.pop(code_field)  # 移除code字段
                    if code_value:
                        # 查找对应的 ID
                        stmt = select(model_class.id).where(
                            and_(
                                model_class.code == code_value,
                                model_class.is_active == True,
                            )
                        )
                        result = self.session.exec(stmt).first()
                        if result:
                            obj_data[id_field] = result
                        else:
                            raise ValidationException(
                                f"找不到 {code_field} 为 '{code_value}' 的记录"
                            )
                    else:
                        # 如果code为空，将对应的ID设为None
                        obj_data[id_field] = None

            # 检查记录的唯一性（排除当前记录）
            path_uniqueness_fields = [
                "category_id",
                "assessment_unit_id",
                "bridge_type_id",
                "part_id",
                "structure_id",
                "component_type_id",
                "component_form_id",
                "disease_id",
                "scale_id",
                "quality_id",
                "quantity_id",
            ]

            # 构建唯一性检查条件，合并现有值和更新值
            uniqueness_conditions = []
            for field in path_uniqueness_fields:
                # 优先使用更新数据中的值，否则使用现有记录的值
                field_value = obj_data.get(field, getattr(db_obj, field, None))
                if field_value is not None:
                    uniqueness_conditions.append(getattr(Paths, field) == field_value)
                else:
                    uniqueness_conditions.append(getattr(Paths, field).is_(None))

            # 检查是否存在完全相同的路径记录（排除当前记录）
            if uniqueness_conditions:
                stmt = select(Paths).where(
                    and_(and_(*uniqueness_conditions), Paths.id != id)
                )
                if hasattr(Paths, "is_active"):
                    stmt = stmt.where(Paths.is_active == True)
                existing_path = self.session.exec(stmt).first()
                if existing_path:
                    raise DuplicateException(
                        resource="Paths",
                        field="path_combination",
                        value="相同的路径组合已存在",
                    )

            # 更新对象
            for field, value in obj_data.items():
                if hasattr(db_obj, field):
                    setattr(db_obj, field, value)

            if hasattr(db_obj, "updated_at"):
                db_obj.updated_at = datetime.now(timezone.utc)

            self.session.commit()
            self.session.refresh(db_obj)

            # 构建返回数据
            response_data = {
                "id": db_obj.id,
                "code": db_obj.code,
                "name": db_obj.name,
            }

            # 添加各个ID对应的关联数据
            response_data.update(self._get_related_data(db_obj))

            return PathsResponse(**response_data)

        except (NotFoundException, DuplicateException, ValidationException):
            self.session.rollback()
            raise
        except Exception as e:
            self.session.rollback()
            raise Exception(f"更新失败: {str(e)}")

    def export_template(self) -> bytes:
        """
        导出路径Excel模板

        Returns:
            Excel文件的字节内容
        """
        try:
            # 创建工作薄
            wb = Workbook()

            # 创建主工作表
            ws_main = wb.active
            ws_main.title = "路径数据"

            # 定义表头
            headers = [
                "编码",
                "名称",
                "桥梁类别",
                "评定单元",
                "桥梁类型",
                "部位",
                "结构类型",
                "部件类型",
                "构件形式",
                "病害类型",
                "标度",
                "定性描述",
                "定量描述",
            ]

            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws_main.cell(row=1, column=col, value=header)
                # 设置表头样式
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                )

            # 添加填写说明
            instruction = "说明：请在下方填写数据，必须与参考数据表中的名称完全一致，编码可留空由系统自动生成"
            ws_main.cell(row=2, column=1, value=instruction)
            ws_main.merge_cells("A2:M2")
            ws_main.cell(row=2, column=1).font = Font(color="FF0000", italic=True)

            # 调整列宽
            for col in range(1, len(headers) + 1):
                ws_main.column_dimensions[
                    ws_main.cell(row=1, column=col).column_letter
                ].width = 15

            # 创建参考数据工作表
            ws_ref = wb.create_sheet(title="参考数据")
            all_options = self.get_options()
            create_reference_data_sheet(ws_ref, all_options)

            # 创建说明工作表
            ws_help = wb.create_sheet(title="填写说明")
            create_help_sheet(ws_help)

            # 保存到字节流
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return buffer.getvalue()

        except Exception as e:
            print(f"导出模板时出错: {e}")
            traceback.print_exc()  # 追踪报错位置
            raise Exception(f"导出模板失败: {str(e)}")

    def import_from_excel(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """
        从Excel导入路径数据

        Args:
            file_content: Excel文件内容
            filename: 文件名

        Returns:
            导入结果报告
        """
        try:
            # 获取参考数据
            all_options = self.get_options()
            reference_data = get_reference_data(all_options)

            # 匹配函数包装器
            def match_func(
                input_name: str, ref_key: str, ref_data: Dict[str, Dict[str, str]]
            ) -> Dict[str, Any]:
                return match_name_to_code(input_name, ref_key, ref_data, self.session)

            # 验证数据
            validation_result = validate_excel_data(
                file_content, filename, reference_data, match_func
            )

            # 统计信息
            total_rows = validation_result["total_rows"]
            valid_rows_count = validation_result["valid_rows_count"]
            invalid_rows_count = validation_result["invalid_rows_count"]

            # 没有任何数据
            if total_rows == 0:
                return {
                    "message": "Excel文件中没有数据",
                    "imported_count": 0,
                    "failed_count": 0,
                    "skipped_count": 0,
                    "total_rows": 0,
                    "validation_result": validation_result,
                }

            # 没有有效数据
            if valid_rows_count == 0:
                return {
                    "message": f"共{total_rows}行数据，但没有有效数据可导入",
                    "imported_count": 0,
                    "failed_count": 0,
                    "skipped_count": total_rows,
                    "total_rows": total_rows,
                    "validation_result": validation_result,
                }

            # 导入有效数据
            valid_rows = validation_result["valid_rows"]
            imported_count = 0
            import_errors = []

            for row_index, row_data in enumerate(valid_rows):
                try:
                    # 创建路径记录
                    path_create = PathsCreate(
                        code=row_data.get("code", ""),
                        name=row_data["name"],
                        category_code=row_data["category_code"],
                        assessment_unit_code=row_data.get("assessment_unit_code"),
                        bridge_type_code=row_data["bridge_type_code"],
                        part_code=row_data["part_code"],
                        structure_code=row_data.get("structure_code"),
                        component_type_code=row_data.get("component_type_code"),
                        component_form_code=row_data.get("component_form_code"),
                        disease_code=row_data["disease_code"],
                        scale_code=row_data["scale_code"],
                        quality_code=row_data.get("quality_code"),
                        quantity_code=row_data.get("quantity_code"),
                    )

                    # 调用创建方法
                    result = self.create(path_create)
                    if result:
                        imported_count += 1
                    else:
                        import_errors.append(
                            {
                                "row": row_index + 4,
                                "error": "创建记录失败",
                            }
                        )

                except Exception as e:
                    import_errors.append(
                        {"row": row_index + 4, "error": f"导入失败: {str(e)}"}
                    )

            failed_count = len(import_errors)
            skipped_count = invalid_rows_count
            message = f"导入完成，共导入 {imported_count} 条数据，共失败 {failed_count} 条，共无效 {skipped_count} 条"

            return {
                "message": message,
                "imported_count": imported_count,
                "failed_count": failed_count,
                "skipped_count": skipped_count,
                "total_rows": total_rows,
                "import_errors": import_errors,
                "validation_result": validation_result,
            }

        except Exception as e:
            print(f"导入Excel数据时出错: {e}")
            traceback.print_exc()
            raise Exception(f"文件处理失败: {str(e)}")


def get_paths_service(session: Session) -> PathsService:
    """
    获取paths服务实例
    """
    return PathsService(session)
