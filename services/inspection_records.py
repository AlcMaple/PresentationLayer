from typing import List, Optional, Dict, Any
from sqlmodel import Session, select, and_
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from fastapi import UploadFile
import traceback
import pandas as pd

from services.file_upload import get_file_upload_service
from models import (
    InspectionRecords,
    Paths,
    UserPaths,
    BridgeTypes,
    BridgeParts,
    BridgeStructures,
    BridgeComponentTypes,
    BridgeComponentForms,
    BridgeDiseases,
    BridgeScales,
    BridgeQualities,
    BridgeQuantities,
    AssessmentUnit,
)
from models.enums import ScalesType
from schemas.inspection_records import (
    InspectionRecordsCreate,
    InspectionRecordsUpdate,
    InspectionRecordsResponse,
    PathValidationRequest,
    FormOptionsResponse,
    DamageTypeOption,
    ScaleOption,
    DamageDetailInfo,
    DamageReferenceRequest,
)
from services.base_crud import BaseCRUDService, PageParams
from exceptions import ValidationException, NotFoundException, SystemException
from utils import (
    get_id_by_code,
    get_damage_type_id_by_name,
    get_scale_id_by_value,
    get_damage_code_by_id,
    get_scale_code_by_id,
)


class InspectionRecordsService(
    BaseCRUDService[InspectionRecords, InspectionRecordsCreate, InspectionRecordsUpdate]
):
    """检查记录服务"""

    def __init__(self, session: Session):
        super().__init__(InspectionRecords, session)

    def _validate_path_exists(self, path_request: PathValidationRequest) -> bool:
        """
        验证路径是否在user_paths表中存在
        """
        try:
            conditions = [
                UserPaths.bridge_instance_name == path_request.bridge_instance_name,
                UserPaths.bridge_type_id == path_request.bridge_type_id,
                UserPaths.part_id == path_request.part_id,
                UserPaths.component_type_id == path_request.component_type_id,
                UserPaths.component_form_id == path_request.component_form_id,
                UserPaths.is_active == True,
            ]

            if path_request.assessment_unit_instance_name:
                conditions.append(
                    UserPaths.assessment_unit_instance_name
                    == path_request.assessment_unit_instance_name
                )
            else:
                conditions.append(UserPaths.assessment_unit_instance_name.is_(None))

            if path_request.structure_id:
                conditions.append(UserPaths.structure_id == path_request.structure_id)
            else:
                conditions.append(UserPaths.structure_id.is_(None))

            # 用户ID验证
            if path_request.user_id is not None:
                conditions.append(UserPaths.user_id == path_request.user_id)
            else:
                # user_id为空代表管理员
                conditions.append(UserPaths.user_id.is_(None))

            stmt = select(UserPaths.id).where(and_(*conditions)).limit(1)
            result = self.session.exec(stmt).first()

            return result is not None

        except Exception as e:
            print(f"验证路径时出错: {e}")
            return False

    def _get_matching_user_path(
        self, path_request: PathValidationRequest
    ) -> Optional[UserPaths]:
        """
        获取匹配的用户路径记录
        """
        try:
            conditions = [
                UserPaths.bridge_instance_name == path_request.bridge_instance_name,
                UserPaths.bridge_type_id == path_request.bridge_type_id,
                UserPaths.part_id == path_request.part_id,
                UserPaths.component_type_id == path_request.component_type_id,
                UserPaths.component_form_id == path_request.component_form_id,
                UserPaths.is_active == True,
            ]

            if path_request.assessment_unit_instance_name:
                conditions.append(
                    UserPaths.assessment_unit_instance_name
                    == path_request.assessment_unit_instance_name
                )
            else:
                conditions.append(UserPaths.assessment_unit_instance_name.is_(None))

            if path_request.structure_id:
                conditions.append(UserPaths.structure_id == path_request.structure_id)
            else:
                conditions.append(UserPaths.structure_id.is_(None))

            # 用户ID验证
            if path_request.user_id is not None:
                conditions.append(UserPaths.user_id == path_request.user_id)
            else:
                conditions.append(UserPaths.user_id.is_(None))

            stmt = select(UserPaths).where(and_(*conditions)).limit(1)
            result = self.session.exec(stmt).first()

            return result

        except Exception as e:
            print(f"获取匹配用户路径时出错: {e}")
            return None

    def _validate_damage_scale_combination(
        self,
        path_request: PathValidationRequest,
        damage_type_code: str,
        scale_code: str,
    ) -> bool:
        """
        验证病害类型和标度的组合是否有效
        通过 user_paths 关联的 paths_id 来验证
        """
        try:

            # 获取匹配的 user_paths 记录
            user_path = self._get_matching_user_path(path_request)
            if not user_path:
                return False

            # 通过 user_paths 的 paths_id 查询对应的 paths 记录
            paths_stmt = select(Paths).where(
                and_(Paths.id == user_path.paths_id, Paths.is_active == True)
            )
            paths_record = self.session.exec(paths_stmt).first()
            if not paths_record:
                return False

            # 通过code获取ID
            damage_type_id = get_id_by_code(
                BridgeDiseases, damage_type_code, self.session
            )
            scale_id = get_id_by_code(BridgeScales, scale_code, self.session)

            if not damage_type_id or not scale_id:
                return False

            # 查询 paths 表中是否存在对应的病害和标度组合，基于 paths_record 的基础路径信息，加上病害和标度信息
            conditions = [
                Paths.category_id == user_path.category_id,
                Paths.bridge_type_id == user_path.bridge_type_id,
                Paths.part_id == user_path.part_id,
                Paths.disease_id == damage_type_id,
                Paths.scale_id == scale_id,
                Paths.is_active == True,
            ]

            if user_path.assessment_unit_id:
                conditions.append(
                    Paths.assessment_unit_id == user_path.assessment_unit_id
                )
            if user_path.component_type_id:
                conditions.append(
                    Paths.component_type_id == user_path.component_type_id
                )
            if user_path.component_form_id:
                conditions.append(
                    Paths.component_form_id == user_path.component_form_id
                )
            if user_path.structure_id:
                conditions.append(Paths.structure_id == user_path.structure_id)

            stmt = select(Paths.id).where(and_(*conditions)).limit(1)
            result = self.session.exec(stmt).first()

            return result is not None

        except Exception as e:
            error_message = f"验证病害标度组合时发生系统错误: {e}"
            raise SystemException(message=error_message, original_error=e)

    async def create(
        self, record_data: InspectionRecordsCreate, image_file: Optional[UploadFile]
    ) -> InspectionRecordsResponse:
        """
        创建检查记录
        """
        try:
            # 验证路径是否存在
            path_request = PathValidationRequest(
                user_id=record_data.user_id,
                bridge_instance_name=record_data.bridge_instance_name,
                assessment_unit_instance_name=record_data.assessment_unit_instance_name,
                bridge_type_id=record_data.bridge_type_id,
                part_id=record_data.part_id,
                structure_id=record_data.structure_id,
                component_type_id=record_data.component_type_id,
                component_form_id=record_data.component_form_id,
            )

            if not self._validate_path_exists(path_request):
                raise ValidationException(
                    "指定的路径组合在系统中不存在 或者 用户没有创建路径数据"
                )

            # 验证病害类型和标度的组合是否有效
            if not self._validate_damage_scale_combination(
                path_request, record_data.damage_type_code, record_data.scale_code
            ):
                raise ValidationException("指定的病害类型和标度组合无效")

            # 通过code获取对应的ID
            damage_type_id = get_id_by_code(
                BridgeDiseases, record_data.damage_type_code, self.session
            )
            scale_id = get_id_by_code(
                BridgeScales, record_data.scale_code, self.session
            )

            if not damage_type_id:
                raise ValidationException(
                    f"找不到病害类型编码: {record_data.damage_type_code}"
                )
            if not scale_id:
                raise ValidationException(f"找不到标度编码: {record_data.scale_code}")

            # 处理图片上传
            image_url = None
            if image_file and image_file.filename:
                file_service = get_file_upload_service()
                success, message, url = await file_service.save_image(image_file)

                if not success:
                    raise ValidationException(f"图片上传失败: {message}")

                image_url = url

            # 创建检查记录
            inspection_record = InspectionRecords(
                user_id=record_data.user_id,
                bridge_instance_name=record_data.bridge_instance_name,
                assessment_unit_instance_name=record_data.assessment_unit_instance_name,
                bridge_type_id=record_data.bridge_type_id,
                part_id=record_data.part_id,
                structure_id=record_data.structure_id,
                component_type_id=record_data.component_type_id,
                component_form_id=record_data.component_form_id,
                damage_type_id=damage_type_id,
                scale_id=scale_id,
                damage_location=record_data.damage_location,
                damage_description=record_data.damage_description,
                image_url=image_url,
                component_name=record_data.component_name,
            )

            self.session.add(inspection_record)
            self.session.commit()
            self.session.refresh(inspection_record)

            return self.get_record_with_details(inspection_record.id)

        except (ValidationException, SystemException) as e:
            self.session.rollback()
            raise e

        except Exception as e:
            self.session.rollback()
            raise SystemException(
                message=f"创建检查记录时发生未知系统错误: {e}", original_error=e
            )

    def get_record_with_details(self, record_id: int) -> InspectionRecordsResponse:
        """
        获取包含详细关联信息的检查记录
        """
        try:
            stmt = select(InspectionRecords).where(
                and_(
                    InspectionRecords.id == record_id,
                    InspectionRecords.is_active == True,
                )
            )
            record = self.session.exec(stmt).first()

            if not record:
                raise NotFoundException(
                    resource="InspectionRecords", identifier=str(record_id)
                )

            # 获取关联信息
            details = self._get_record_related_data(record)

            # 获取表单选项
            path_request = PathValidationRequest(
                user_id=record.user_id,
                bridge_instance_name=record.bridge_instance_name,
                assessment_unit_instance_name=record.assessment_unit_instance_name,
                bridge_type_id=record.bridge_type_id,
                part_id=record.part_id,
                structure_id=record.structure_id,
                component_type_id=record.component_type_id,
                component_form_id=record.component_form_id,
            )
            form_options = self.get_form_options_by_path(path_request)

            return InspectionRecordsResponse(
                id=record.id,
                user_id=record.user_id,
                bridge_instance_name=record.bridge_instance_name,
                assessment_unit_instance_name=record.assessment_unit_instance_name,
                bridge_type_id=record.bridge_type_id,
                bridge_type_code=details.get("bridge_type_code"),
                bridge_type_name=details.get("bridge_type_name"),
                part_id=record.part_id,
                part_code=details.get("part_code"),
                part_name=details.get("part_name"),
                structure_id=record.structure_id,
                structure_code=details.get("structure_code"),
                structure_name=details.get("structure_name"),
                component_type_id=record.component_type_id,
                component_type_code=details.get("component_type_code"),
                component_type_name=details.get("component_type_name"),
                component_form_id=record.component_form_id,
                component_form_code=details.get("component_form_code"),
                component_form_name=details.get("component_form_name"),
                damage_type_id=record.damage_type_id,
                damage_type_code=details.get("damage_type_code"),
                damage_type_name=details.get("damage_type_name"),
                scale_id=record.scale_id,
                scale_code=details.get("scale_code"),
                scale_name=details.get("scale_name"),
                scale_value=details.get("scale_value"),
                damage_location=record.damage_location,
                damage_description=record.damage_description,
                image_url=record.image_url,
                created_at=record.created_at,
                updated_at=record.updated_at,
                is_active=record.is_active,
                form_options=form_options,
                component_name=record.component_name,
            )

        except NotFoundException:
            raise
        except Exception as e:
            raise Exception(f"获取检查记录详情失败: {str(e)}")

    def _get_record_related_data(self, record: InspectionRecords) -> Dict[str, Any]:
        """
        获取检查记录的关联数据
        """
        details = {}

        # 定义查询映射
        field_mappings = [
            ("bridge_type_id", BridgeTypes, "bridge_type"),
            ("part_id", BridgeParts, "part"),
            ("structure_id", BridgeStructures, "structure"),
            ("component_type_id", BridgeComponentTypes, "component_type"),
            ("component_form_id", BridgeComponentForms, "component_form"),
        ]

        for field_name, model_class, prefix in field_mappings:
            field_id = getattr(record, field_name, None)
            if field_id:
                try:
                    stmt = select(model_class.code, model_class.name).where(
                        model_class.id == field_id
                    )
                    result = self.session.exec(stmt).first()
                    if result:
                        details[f"{prefix}_code"] = result[0]
                        details[f"{prefix}_name"] = result[1]
                    else:
                        details[f"{prefix}_code"] = None
                        details[f"{prefix}_name"] = None
                except:
                    details[f"{prefix}_code"] = None
                    details[f"{prefix}_name"] = None
            else:
                details[f"{prefix}_code"] = None
                details[f"{prefix}_name"] = None

        # 获取病害类型信息
        try:
            stmt = select(BridgeDiseases.code, BridgeDiseases.name).where(
                BridgeDiseases.id == record.damage_type_id
            )
            disease_result = self.session.exec(stmt).first()
            if disease_result:
                details["damage_type_code"] = disease_result[0]
                details["damage_type_name"] = disease_result[1]
            else:
                details["damage_type_code"] = None
                details["damage_type_name"] = None
        except:
            details["damage_type_code"] = None
            details["damage_type_name"] = None

        # 获取标度信息
        try:
            stmt = select(
                BridgeScales.code,
                BridgeScales.name,
                BridgeScales.scale_type,
                BridgeScales.scale_value,
                BridgeScales.min_value,
                BridgeScales.max_value,
                BridgeScales.unit,
                BridgeScales.display_text,
            ).where(BridgeScales.id == record.scale_id)
            scale_result = self.session.exec(stmt).first()
            if scale_result:
                (
                    code,
                    name,
                    scale_type,
                    scale_value,
                    min_value,
                    max_value,
                    unit,
                    display_text,
                ) = scale_result

                details["scale_code"] = code

                # 根据标度类型构建显示名称
                if scale_type == ScalesType.NUMERIC:
                    display_name = str(scale_value) if scale_value is not None else name
                    details["scale_value"] = scale_value
                elif scale_type == ScalesType.RANGE:
                    if min_value is not None and max_value is not None and unit:
                        display_name = f"{min_value}-{max_value}{unit}"
                    else:
                        display_name = name
                    details["scale_value"] = None
                elif scale_type == ScalesType.TEXT:
                    display_name = display_text if display_text else name
                    details["scale_value"] = None
                else:
                    display_name = name
                    details["scale_value"] = scale_value

                details["scale_name"] = display_name
            else:
                details["scale_code"] = None
                details["scale_name"] = None
                details["scale_value"] = None
        except:
            details["scale_code"] = None
            details["scale_name"] = None
            details["scale_value"] = None

        return details

    async def update(
        self,
        record_id: int,
        update_data: InspectionRecordsUpdate,
        image_file: Optional[UploadFile],
    ) -> InspectionRecordsResponse:
        """
        更新检查记录
        """
        try:
            # 查询现有记录
            existing_record = self.get_by_id(record_id)
            if not existing_record:
                raise NotFoundException(
                    resource="InspectionRecords", identifier=str(record_id)
                )

            # 如果更新了病害类型或标度，验证组合的有效性
            if update_data.damage_type_code or update_data.scale_code:
                path_request = PathValidationRequest(
                    user_id=existing_record.user_id,
                    bridge_instance_name=existing_record.bridge_instance_name,
                    assessment_unit_instance_name=existing_record.assessment_unit_instance_name,
                    bridge_type_id=existing_record.bridge_type_id,
                    part_id=existing_record.part_id,
                    structure_id=existing_record.structure_id,
                    component_type_id=existing_record.component_type_id,
                    component_form_id=existing_record.component_form_id,
                )

                # 病害类型和标度编码
                damage_code = update_data.damage_type_code
                scale_code = update_data.scale_code

                # 如果没有提供新的code，使用现有记录的code
                if not damage_code:
                    # 通过现有记录的ID获取code
                    stmt = select(BridgeDiseases.code).where(
                        BridgeDiseases.id == existing_record.damage_type_id
                    )
                    damage_code = self.session.exec(stmt).first()

                if not scale_code:
                    stmt = select(BridgeScales.code).where(
                        BridgeScales.id == existing_record.scale_id
                    )
                    scale_code = self.session.exec(stmt).first()

                # 验证病害类型和标度的组合
                if not self._validate_damage_scale_combination(
                    path_request, damage_code, scale_code
                ):
                    raise ValidationException("指定的病害类型和标度组合无效")

            # 更新字段
            if update_data.damage_type_code is not None:
                damage_type_id = get_id_by_code(
                    BridgeDiseases, update_data.damage_type_code, self.session
                )
                if not damage_type_id:
                    raise ValidationException(
                        f"找不到病害类型编码: {update_data.damage_type_code}"
                    )
                existing_record.damage_type_id = damage_type_id

            if update_data.scale_code is not None:
                scale_id = get_id_by_code(
                    BridgeScales, update_data.scale_code, self.session
                )
                if not scale_id:
                    raise ValidationException(
                        f"找不到标度编码: {update_data.scale_code}"
                    )
                existing_record.scale_id = scale_id

            if update_data.damage_location is not None:
                existing_record.damage_location = update_data.damage_location
            if update_data.damage_description is not None:
                existing_record.damage_description = update_data.damage_description
            if update_data.image_url is not None:
                existing_record.image_url = update_data.image_url
            if update_data.component_name is not None:
                existing_record.component_name = update_data.component_name
            if image_file and image_file.filename:
                file_service = get_file_upload_service()
                success, message, url = await file_service.save_image(image_file)

                if not success:
                    raise ValidationException(f"图片上传失败: {message}")

                # 更新图片URL
                existing_record.image_url = url

            # 更新时间
            existing_record.updated_at = datetime.now(timezone.utc)

            # 保存更改
            self.session.commit()
            self.session.refresh(existing_record)

            # 返回详细信息
            return self.get_record_with_details(record_id)

        except (NotFoundException, ValidationException):
            self.session.rollback()
            raise
        except Exception as e:
            self.session.rollback()
            raise Exception(f"更新检查记录失败: {str(e)}")

    def _get_default_id(self, model_class) -> Optional[int]:
        """获取名称为"-"的记录ID"""
        try:
            stmt = (
                select(model_class.id)
                .where(and_(model_class.name == "-", model_class.is_active == True))
                .limit(1)
            )
            result = self.session.exec(stmt).first()
            return result
        except Exception as e:
            print(f"获取默认ID时出错 {model_class.__name__}: {e}")
            return None

    def get_form_options_by_path(
        self, path_request: PathValidationRequest
    ) -> FormOptionsResponse:
        """
        根据路径获取表单选项
        """
        try:
            # 获取用户路径记录
            user_path = self._get_matching_user_path(path_request)
            if not user_path:
                return FormOptionsResponse(damage_types=[], scales_by_damage={})

            # 通过用户路径的 paths_id 获取基础路径信息
            paths_stmt = select(Paths).where(
                and_(Paths.id == user_path.paths_id, Paths.is_active == True)
            )
            paths_record = self.session.exec(paths_stmt).first()
            if not paths_record:
                return FormOptionsResponse(damage_types=[], scales_by_damage={})

            assessment_unit_id = paths_record.assessment_unit_id
            if assessment_unit_id is None:
                assessment_unit_id = self._get_default_id(AssessmentUnit)

            structure_id = paths_record.structure_id
            if structure_id is None:
                structure_id = self._get_default_id(BridgeStructures)

            component_type_id = paths_record.component_type_id
            if component_type_id is None:
                component_type_id = self._get_default_id(BridgeComponentTypes)

            component_form_id = paths_record.component_form_id
            if component_form_id is None:
                component_form_id = self._get_default_id(BridgeComponentForms)

            base_conditions = [
                Paths.category_id == paths_record.category_id,
                Paths.assessment_unit_id == paths_record.assessment_unit_id,
                Paths.bridge_type_id == paths_record.bridge_type_id,
                Paths.part_id == paths_record.part_id,
                Paths.structure_id == paths_record.structure_id,
                Paths.component_type_id == paths_record.component_type_id,
                Paths.component_form_id == paths_record.component_form_id,
                Paths.is_active == True,
            ]

            # 查询符合条件的paths病害和标度记录
            paths_stmt = (
                select(Paths.disease_id, Paths.scale_id)
                .where(
                    and_(
                        *base_conditions,
                        Paths.disease_id.is_not(None),
                        Paths.scale_id.is_not(None),
                    )
                )
                .distinct()
            )

            path_results = self.session.exec(paths_stmt).all()

            if not path_results:
                return FormOptionsResponse(damage_types=[], scales_by_damage={})

            # 收集所有的disease_id和scale_id
            disease_ids = set()
            scale_ids = set()
            disease_scale_pairs = set()

            for result in path_results:
                disease_ids.add(result.disease_id)
                scale_ids.add(result.scale_id)
                disease_scale_pairs.add((result.disease_id, result.scale_id))

            # 查询病害类型信息
            diseases_stmt = (
                select(BridgeDiseases.id, BridgeDiseases.code, BridgeDiseases.name)
                .where(
                    and_(
                        BridgeDiseases.id.in_(disease_ids),
                        BridgeDiseases.is_active == True,
                    )
                )
                .order_by(BridgeDiseases.code)
            )

            disease_results = self.session.exec(diseases_stmt).all()
            disease_map = {
                r.id: {"code": r.code, "name": r.name} for r in disease_results
            }

            # 查询标度信息
            scales_stmt = (
                select(
                    BridgeScales.id,
                    BridgeScales.code,
                    BridgeScales.scale_type,
                    BridgeScales.scale_value,
                    BridgeScales.min_value,
                    BridgeScales.max_value,
                    BridgeScales.unit,
                    BridgeScales.display_text,
                )
                .where(
                    and_(BridgeScales.id.in_(scale_ids), BridgeScales.is_active == True)
                )
                .order_by(BridgeScales.scale_value)
            )

            scale_results = self.session.exec(scales_stmt).all()
            scale_map = {}

            for r in scale_results:
                # 标度显示名称
                if r.scale_type == ScalesType.NUMERIC:
                    display_name = (
                        str(r.scale_value) if r.scale_value is not None else r.code
                    )
                elif r.scale_type == ScalesType.RANGE:
                    if r.min_value is not None and r.max_value is not None and r.unit:
                        display_name = f"{r.min_value}-{r.max_value}{r.unit}"
                    else:
                        display_name = r.code
                elif r.scale_type == ScalesType.TEXT:
                    display_name = r.display_text if r.display_text else r.code
                else:
                    display_name = r.code

                scale_map[r.id] = {
                    "code": r.code,
                    "name": display_name,
                    "value": r.scale_value,
                }

            # 按病害类型分组
            damage_scale_map = {}

            for disease_id, scale_id in disease_scale_pairs:
                if disease_id not in disease_map or scale_id not in scale_map:
                    continue

                disease_info = disease_map[disease_id]
                scale_info = scale_map[scale_id]

                disease_code = disease_info["code"]

                if disease_code not in damage_scale_map:
                    damage_scale_map[disease_code] = {
                        "damage_type": DamageTypeOption(
                            code=disease_code,
                            name=disease_info["name"],
                        ),
                        "scales": [],
                    }

                # 避免重复的标度
                scale_option = ScaleOption(
                    code=scale_info["code"],
                    name=scale_info["name"],
                    value=scale_info["value"],
                )

                existing_codes = [
                    s.code for s in damage_scale_map[disease_code]["scales"]
                ]
                if scale_option.code not in existing_codes:
                    damage_scale_map[disease_code]["scales"].append(scale_option)

            return FormOptionsResponse(
                damage_types=[
                    item["damage_type"] for item in damage_scale_map.values()
                ],
                scales_by_damage={k: v["scales"] for k, v in damage_scale_map.items()},
            )

        except Exception as e:
            print(f"获取表单选项时出错: {e}")
            traceback.print_exc()
            return FormOptionsResponse(damage_types=[], scales_by_damage={})

    def get_damage_reference_info(
        self, request: DamageReferenceRequest
    ) -> DamageDetailInfo:
        """
        根据路径和病害类型获取参考信息
        """
        try:

            # 获取匹配的用户路径记录
            path_request = PathValidationRequest(
                bridge_instance_name=request.bridge_instance_name,
                assessment_unit_instance_name=request.assessment_unit_instance_name,
                bridge_type_id=request.bridge_type_id,
                part_id=request.part_id,
                structure_id=request.structure_id,
                component_type_id=request.component_type_id,
                component_form_id=request.component_form_id,
            )

            user_path = self._get_matching_user_path(path_request)
            if not user_path:
                return DamageDetailInfo(
                    damage_type_code=request.damage_type_code,
                    damage_type_name="",
                    scales=[],
                    qualities=[],
                    quantities=[],
                )

            # 通过用户路径的 paths_id 获取基础路径信息
            paths_stmt = select(Paths).where(
                and_(Paths.id == user_path.paths_id, Paths.is_active == True)
            )
            paths_record = self.session.exec(paths_stmt).first()
            if not paths_record:
                return DamageDetailInfo(
                    damage_type_code=request.damage_type_code,
                    damage_type_name="",
                    scales=[],
                    qualities=[],
                    quantities=[],
                )

            # 通过code获取病害类型ID
            damage_type_id = get_id_by_code(
                BridgeDiseases, request.damage_type_code, self.session
            )
            if not damage_type_id:
                return DamageDetailInfo(
                    damage_type_code=request.damage_type_code,
                    damage_type_name="",
                    scales=[],
                    qualities=[],
                    quantities=[],
                )

            # 查询条件
            base_conditions = [
                Paths.category_id == paths_record.category_id,
                Paths.assessment_unit_id == paths_record.assessment_unit_id,
                Paths.bridge_type_id == paths_record.bridge_type_id,
                Paths.part_id == paths_record.part_id,
                Paths.structure_id == paths_record.structure_id,
                Paths.component_type_id == paths_record.component_type_id,
                Paths.component_form_id == paths_record.component_form_id,
                Paths.disease_id == damage_type_id,
                Paths.is_active == True,
            ]

            # 查询该病害类型下的所有paths记录
            paths_stmt = (
                select(Paths.scale_id, Paths.quality_id, Paths.quantity_id)
                .where(and_(*base_conditions))
                .distinct()
            )

            path_results = self.session.exec(paths_stmt).all()

            # 获取病害类型名称
            disease_stmt = select(BridgeDiseases.name).where(
                BridgeDiseases.id == damage_type_id
            )
            disease_name = self.session.exec(disease_stmt).first() or ""

            if not path_results:
                return DamageDetailInfo(
                    damage_type_code=request.damage_type_code,
                    damage_type_name=disease_name,
                    scales=[],
                    qualities=[],
                    quantities=[],
                )

            # 收集ID
            scale_ids = []
            quality_ids = []
            quantity_ids = []

            for result in path_results:
                if result.scale_id:
                    scale_ids.append(result.scale_id)
                if result.quality_id:
                    quality_ids.append(result.quality_id)
                if result.quantity_id:
                    quantity_ids.append(result.quantity_id)

            # 查询标度信息
            scales = []
            if scale_ids:
                scales_stmt = (
                    select(
                        BridgeScales.code,
                        BridgeScales.scale_type,
                        BridgeScales.scale_value,
                        BridgeScales.min_value,
                        BridgeScales.max_value,
                        BridgeScales.unit,
                        BridgeScales.display_text,
                    )
                    .where(
                        and_(
                            BridgeScales.id.in_(scale_ids),
                            BridgeScales.is_active == True,
                        )
                    )
                    .order_by(BridgeScales.scale_value)
                )

                scale_results = self.session.exec(scales_stmt).all()

                for r in scale_results:
                    # 构建标度显示名称
                    if r.scale_type == ScalesType.NUMERIC:
                        display_name = (
                            str(r.scale_value) if r.scale_value is not None else r.code
                        )
                    elif r.scale_type == ScalesType.RANGE:
                        if (
                            r.min_value is not None
                            and r.max_value is not None
                            and r.unit
                        ):
                            display_name = f"{r.min_value}-{r.max_value}{r.unit}"
                        else:
                            display_name = r.code
                    elif r.scale_type == ScalesType.TEXT:
                        display_name = r.display_text if r.display_text else r.code
                    else:
                        display_name = r.code

                    scales.append(
                        ScaleOption(code=r.code, name=display_name, value=r.scale_value)
                    )

            # 查询定性描述
            qualities = []
            for quality_id in quality_ids:
                quality_stmt = select(BridgeQualities.name).where(
                    and_(
                        BridgeQualities.id == quality_id,
                        BridgeQualities.is_active == True,
                    )
                )
                result = self.session.exec(quality_stmt).first()
                if result:
                    qualities.append(result)

            # 查询定量描述
            quantities = []
            for quantity_id in quantity_ids:
                quantity_stmt = select(BridgeQuantities.name).where(
                    and_(
                        BridgeQuantities.id == quantity_id,
                        BridgeQuantities.is_active == True,
                    )
                )
                result = self.session.exec(quantity_stmt).first()
                if result:
                    quantities.append(result)

            return DamageDetailInfo(
                damage_type_code=request.damage_type_code,
                damage_type_name=disease_name,
                scales=scales,
                qualities=qualities,
                quantities=quantities,
            )

        except Exception as e:
            print(f"获取病害参考信息时出错: {e}")
            traceback.print_exc()
            return DamageDetailInfo(
                damage_type_code=request.damage_type_code,
                damage_type_name="",
                scales=[],
                qualities=[],
                quantities=[],
            )

    def export_template(self, path_request: PathValidationRequest) -> bytes:
        """
        导出检查记录Excel模板

        Args:
            path_request: 路径验证请求

        Returns:
            Excel文件的字节内容
        """
        try:

            # 验证路径是否存在
            if not self._validate_path_exists(path_request):
                raise ValidationException("指定的路径组合在系统中不存在")

            # 获取该路径下的表单选项
            form_options = self.get_form_options_by_path(path_request)

            # 创建工作薄
            wb = Workbook()

            # 创建主工作表
            ws_main = wb.active
            ws_main.title = "检查记录数据"

            # 定义表头
            headers = [
                "构件名称",
                "病害类型",
                "标度值",
                "病害位置",
                "病害程度",
            ]

            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws_main.cell(row=1, column=col, value=header)
                # 设置表头样式
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                )

            # 添加填写说明
            instruction = "说明：构件名称、病害位置、病害程度可选；病害类型和标度值请参考对应的参考数据表"
            ws_main.cell(row=2, column=1, value=instruction)
            ws_main.merge_cells("A2:F2")
            ws_main.cell(row=2, column=1).font = Font(color="FF0000", italic=True)

            # 调整列宽
            column_widths = [15, 20, 15, 25, 25, 30]
            for col, width in enumerate(column_widths, 1):
                ws_main.column_dimensions[
                    ws_main.cell(row=1, column=col).column_letter
                ].width = width

            # 创建病害类型参考数据表
            self._create_damage_reference_sheet(wb, form_options)

            # 创建使用说明表
            self._create_help_sheet_for_inspection(wb)

            # 保存到字节流
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return buffer.getvalue()

        except ValidationException:
            raise
        except Exception as e:
            print(f"导出检查记录模板时出错: {e}")
            traceback.print_exc()
            raise Exception(f"导出模板失败: {str(e)}")

    def _create_damage_reference_sheet(self, wb, form_options: FormOptionsResponse):
        """创建病害类型参考数据工作表"""
        try:
            ws_ref = wb.create_sheet(title="病害参考数据")

            # 写入列标题
            headers = ["病害类型名称", "标度名称"]
            for col, header in enumerate(headers, 1):
                cell = ws_ref.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"
                )

            # 写入数据
            row = 2
            for damage_type in form_options.damage_types:
                damage_code = damage_type.code
                damage_name = damage_type.name

                # 获取该病害类型对应的标度选项
                scales = form_options.scales_by_damage.get(damage_code, [])

                if scales:
                    for scale in scales:
                        ws_ref.cell(row=row, column=1, value=damage_name)
                        ws_ref.cell(row=row, column=2, value=scale.name)
                        row += 1

            # 调整列宽
            for col in range(1, 3):
                ws_ref.column_dimensions[
                    ws_ref.cell(row=1, column=col).column_letter
                ].width = 20

        except Exception as e:
            print(f"创建病害参考数据表时出错: {e}")

    def _create_help_sheet_for_inspection(self, wb):
        """创建检查记录使用说明工作表"""
        try:
            ws_help = wb.create_sheet(title="填写说明")
            help_content = [
                ["检查记录数据导入模板使用说明", ""],
                ["", ""],
                ["1. 数据填写要求", ""],
                ["• 构件名称：可自由填写，用于标识具体构件", ""],
                ["• 病害类型：必须从'病害参考数据'表中选择对应的名称", ""],
                ["• 标度值：必须与病害类型匹配，参考'病害参考数据'表", ""],
                ["• 病害位置：可自由填写，描述病害的具体位置", ""],
                ["• 病害程度：可自由填写，描述病害的详细情况", ""],
                ["", ""],
                ["2. 注意事项", ""],
                ["• 病害类型和标度值必须匹配，系统会验证组合的有效性", ""],
                ["• 建议使用复制粘贴方式从参考数据表中选择", ""],
                ["• 所有数据都可以为空，但病害类型和标度值必须必须填写", ""],
                ["", ""],
                ["3. 导入流程", ""],
                ["• 填写完数据后保存Excel文件", ""],
                ["• 在系统中选择相同的路径", ""],
                ["• 上传Excel文件进行导入", ""],
                ["• 系统会验证数据并生成导入报告", ""],
            ]

            for row, (content1, content2) in enumerate(help_content, 1):
                ws_help.cell(row=row, column=1, value=content1)
                ws_help.cell(row=row, column=2, value=content2)

                # 设置样式
                if "说明" in content1:
                    ws_help.cell(row=row, column=1).font = Font(bold=True, size=14)
                elif any(
                    keyword in content1
                    for keyword in ["路径信息", "填写要求", "注意事项", "导入流程"]
                ):
                    ws_help.cell(row=row, column=1).font = Font(
                        bold=True, color="0066CC"
                    )

            # 调整列宽
            ws_help.column_dimensions["A"].width = 40
            ws_help.column_dimensions["B"].width = 50

        except Exception as e:
            print(f"创建使用说明工作表时出错: {e}")

    async def import_from_excel(
        self, file_content: bytes, path_request: PathValidationRequest, filename: str
    ) -> Dict[str, Any]:
        """
        从 Excel 文件中导入记录数据

        Args:
            file_content: Excel 文件内容
            path_request: 路径验证请求
            filename: 文件名

        Returns:
            导入结果报告
        """
        try:
            # 验证路径是否存在
            if not self._validate_path_exists(path_request):
                raise ValidationException("指定的路径组合在系统中不存在")

            # 读取Excel文件
            buffer = BytesIO(file_content)
            df = pd.read_excel(buffer, sheet_name="检查记录数据", header=0)
            df = df.iloc[1:].reset_index(drop=True)  # 去除说明行

            # 删除空白行
            df = df.dropna(how="all").reset_index(drop=True)

            print(f"读取到 {len(df)} 行数据")

            if len(df) == 0:
                return {
                    "message": "Excel文件中没有数据",
                    "imported_count": 0,
                    "failed_count": 0,
                    "total_rows": 0,
                    "errors": [],
                }

            # 验证表头
            expected_columns = [
                "构件名称",
                "病害类型",
                "标度值",
                "病害位置",
                "病害程度",
            ]
            if list(df.columns[:5]) != expected_columns:
                raise ValidationException("Excel文件表头格式不正确")

            # 逐行验证和处理数据
            imported_count = 0
            errors = []

            for index, row in df.iterrows():
                row_number = index + 3

                try:
                    # 验证病害类型和标度值
                    damage_type_name = str(row.get("病害类型", "")).strip()
                    scale_value_str = str(row.get("标度值", "")).strip()

                    if not damage_type_name or damage_type_name == "nan":
                        errors.append({"row": row_number, "error": "病害类型不能为空"})
                        continue

                    if not scale_value_str or scale_value_str == "nan":
                        errors.append({"row": row_number, "error": "标度值不能为空"})
                        continue

                    # 通过病害类型名称查找ID
                    damage_type_id = get_damage_type_id_by_name(
                        damage_type_name, self.session
                    )
                    if not damage_type_id:
                        errors.append(
                            {
                                "row": row_number,
                                "error": f"找不到病害类型: {damage_type_name}",
                            }
                        )
                        continue

                    # 通过标度值查找ID
                    scale_id = get_scale_id_by_value(scale_value_str, self.session)
                    if not scale_id:
                        errors.append(
                            {
                                "row": row_number,
                                "error": f"找不到标度值: {scale_value_str}",
                            }
                        )
                        continue

                    # 验证病害类型和标度的组合是否有效
                    if not self._validate_damage_scale_combination(
                        path_request,
                        get_damage_code_by_id(damage_type_id, self.session),
                        get_scale_code_by_id(scale_id, self.session),
                    ):
                        errors.append(
                            {
                                "row": row_number,
                                "error": f"病害类型 {damage_type_name} 和标度值 {scale_value_str} 的组合无效",
                            }
                        )
                        continue

                    # 创建检查记录
                    inspection_record = InspectionRecords(
                        user_id=path_request.user_id,
                        bridge_instance_name=path_request.bridge_instance_name,
                        assessment_unit_instance_name=path_request.assessment_unit_instance_name,
                        bridge_type_id=path_request.bridge_type_id,
                        part_id=path_request.part_id,
                        structure_id=path_request.structure_id,
                        component_type_id=path_request.component_type_id,
                        component_form_id=path_request.component_form_id,
                        damage_type_id=damage_type_id,
                        scale_id=scale_id,
                        component_name=str(row.get("构件名称", "")).strip() or None,
                        damage_location=str(row.get("病害位置", "")).strip() or None,
                        damage_description=str(row.get("病害程度", "")).strip() or None,
                    )

                    self.session.add(inspection_record)
                    self.session.commit()
                    imported_count += 1
                except Exception as e:
                    errors.append({"row": row_number, "error": f"处理失败: {str(e)}"})

            failed_count = len(errors)
            message = (
                f"导入完成，共导入 {imported_count} 条记录，失败 {failed_count} 条"
            )

            return {
                "message": message,
                "imported_count": imported_count,
                "failed_count": failed_count,
                "total_rows": len(df),
                "errors": errors,
            }
        except ValidationException:
            raise
        except Exception as e:
            print(f"导入检查记录数据时出错: {e}")
            traceback.print_exc()
            raise Exception(f"导入检查记录数据失败: {str(e)}")

    def get_filter_options(
        self, user_id: Optional[int] = None
    ) -> Dict[str, List[Dict[str, Any]]]:
        """
        获取检查记录分页查询的过滤选项

        Returns:
            包含各层级选项的字典
        """
        try:
            options = {}

            # 定义字段映射
            table_configs = [
                ("parts", BridgeParts, "part", InspectionRecords.part_id),
                (
                    "structures",
                    BridgeStructures,
                    "structure",
                    InspectionRecords.structure_id,
                ),
                (
                    "component_types",
                    BridgeComponentTypes,
                    "component_type",
                    InspectionRecords.component_type_id,
                ),
                (
                    "component_forms",
                    BridgeComponentForms,
                    "component_form",
                    InspectionRecords.component_form_id,
                ),
            ]

            for table_name, model_class, option_key, record_field in table_configs:
                try:
                    base_conditions = [
                        record_field.is_not(None),
                        InspectionRecords.is_active == True,
                    ]

                    # 用户ID过滤
                    if user_id is not None:
                        base_conditions.append(InspectionRecords.user_id == user_id)
                    else:
                        base_conditions.append(InspectionRecords.user_id.is_(None))

                    # 查询在inspection_records表中存在的ID
                    existing_ids_stmt = (
                        select(record_field).where(and_(*base_conditions)).distinct()
                    )
                    existing_ids = self.session.exec(existing_ids_stmt).all()

                    if existing_ids:
                        # 查询对应的名称信息
                        stmt = (
                            select(model_class.id, model_class.name)
                            .where(
                                and_(
                                    model_class.id.in_(existing_ids),
                                    model_class.is_active == True,
                                )
                            )
                            .order_by(model_class.name)
                        )

                        results = self.session.exec(stmt).all()
                        options[option_key] = [
                            {"id": row[0], "name": row[1]} for row in results
                        ]
                    else:
                        options[option_key] = []

                except Exception as e:
                    print(f"查询表 {table_name} 选项时出错: {e}")
                    options[option_key] = []

            return options

        except Exception as e:
            print(f"获取过滤选项时出错: {e}")
            return {}


def get_inspection_records_service(session: Session) -> InspectionRecordsService:
    """获取检查记录服务实例"""
    return InspectionRecordsService(session)
