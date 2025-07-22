from typing import Dict, Any
from io import BytesIO
from openpyxl.styles import Font, PatternFill
import traceback
import pandas as pd


def create_reference_data_sheet(ws_ref, all_options: Dict[str, Any]):
    """创建参考数据工作表"""
    try:
        # 定义参考表的列
        ref_columns = [
            ("桥梁类别", "categories"),
            ("评定单元", "assessment_units"),
            ("桥梁类型", "bridge_types"),
            ("部位", "bridge_parts"),
            ("结构类型", "bridge_structures"),
            ("部件类型", "bridge_component_types"),
            ("构件形式", "bridge_component_forms"),
            ("病害类型", "bridge_diseases"),
            ("标度", "bridge_scales"),
            ("定性描述", "bridge_qualities"),
            ("定量描述", "bridge_quantities"),
        ]

        # 写入列标题
        for col, (title, _) in enumerate(ref_columns, 1):
            cell = ws_ref.cell(row=1, column=col, value=title)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"
            )

        # 写入数据
        for col, (title, option_key) in enumerate(ref_columns, 1):
            if option_key in all_options:
                options = all_options[option_key]
                for row, option in enumerate(options, 2):  # 从第2行开始
                    name = option.get("name", "")
                    ws_ref.cell(row=row, column=col, value=name)

            # 调整列宽
            ws_ref.column_dimensions[
                ws_ref.cell(row=1, column=col).column_letter
            ].width = 20

    except Exception as e:
        print(f"创建参考数据表时出错: {e}")


def create_help_sheet(ws_help):
    """创建说明工作表"""
    try:
        help_content = [
            ["桥梁路径数据导入模板使用说明", ""],
            ["", ""],
            ["1. 基本要求", ""],
            ["• 编码列：可留空，系统自动生成", ""],
            ["• 名称列：必填，用于标识整条路径", ""],
            ["• 其他列：请填写与参考数据表完全一致的名称", ""],
            ["", ""],
            ["2. 数据填写", ""],
            ["• 打开'参考数据'工作表查看所有可用选项", ""],
            ["• 复制粘贴参考数据中的名称，确保完全一致", ""],
            ["• 注意大小写和空格", ""],
            ["", ""],
            ["3. 导入规则", ""],
            ["• 系统会严格匹配名称", ""],
            ["• 无法匹配的行将被跳过", ""],
            ["• 导入后会生成详细报告", ""],
            ["", ""],
            ["4. 常见问题", ""],
            ["• 名称拼写错误 → 检查参考数据表", ""],
            ["• 多余的空格 → 使用参考数据表复制粘贴", ""],
            ["• 大小写不匹配 → 严格按照参考数据填写", ""],
            ["", ""],
            ["5. 建议操作", ""],
            ["• 先填写少量数据测试", ""],
            ["• 使用复制粘贴避免输入错误", ""],
            ["• 保存备份以便修改", ""],
        ]

        for row, (content1, content2) in enumerate(help_content, 1):
            ws_help.cell(row=row, column=1, value=content1)
            ws_help.cell(row=row, column=2, value=content2)

            # 设置标题样式
            if "说明" in content1:
                ws_help.cell(row=row, column=1).font = Font(bold=True, size=14)
            elif (
                content1.endswith("要求")
                or content1.endswith("填写")
                or content1.endswith("规则")
                or content1.endswith("问题")
                or content1.endswith("操作")
            ):
                ws_help.cell(row=row, column=1).font = Font(bold=True, color="0066CC")

        # 调整列宽
        ws_help.column_dimensions["A"].width = 30
        ws_help.column_dimensions["B"].width = 50

    except Exception as e:
        print(f"创建说明工作表时出错: {e}")


def validate_excel_data(
    file_content: bytes,
    filename: str,
    reference_data: Dict[str, Dict[str, str]],
    match_name_to_code_func,
) -> Dict[str, Any]:
    """
    验证 Excel 数据

    Args:
        file_content: Excel文件内容
        filename: 文件名
        reference_data: 参考数据字典
        match_name_to_code_func: 匹配名称到编码的函数

    Returns:
        验证结果报告
    """
    try:
        # 读取Excel文件
        buffer = BytesIO(file_content)

        # 读取主数据工作表
        df = pd.read_excel(buffer, sheet_name="路径数据", header=0)
        df = df.iloc[1:].reset_index(drop=True)  # 去除标题行

        # 删除空白行
        df = df.dropna(how="all").reset_index(drop=True)

        # 定义验证结果
        validation_results = {
            "filename": filename,
            "total_rows": len(df),
            "valid_rows": [],
            "invalid_rows": [],
            "errors": [],
            "valid_rows_count": 0,
            "invalid_rows_count": 0,
        }

        # 逐行验证
        for index, row in df.iterrows():
            row_validation = validate_row(
                row, index + 3, reference_data, match_name_to_code_func
            )

            if row_validation["is_valid"]:
                validation_results["valid_rows"].append(row_validation["data"])
                validation_results["valid_rows_count"] += 1
            else:
                validation_results["invalid_rows"].append(row_validation)
                validation_results["invalid_rows_count"] += 1

        return validation_results

    except Exception as e:
        print(f"验证Excel数据时出错: {e}")
        traceback.print_exc()
        raise


def validate_row(
    row: pd.Series,
    row_number: int,
    reference_data: Dict[str, Dict[str, str]],
    match_name_to_code_func,
) -> Dict[str, Any]:
    """
    验证单行数据

    Args:
        row: 数据行
        row_number: 行号
        reference_data: 参考数据
        match_name_to_code_func: 匹配名称到编码的函数

    Returns:
        验证结果字典，包含是否有效和错误信息
    """
    result = {
        "row_number": row_number,
        "is_valid": True,
        "errors": [],
        "warnings": [],
        "data": {},
    }

    try:
        column_mapping = {
            # 列名: (内部字段名, 参考数据中的键名, 是否为必填项)
            "编码": ("code", None, False),
            "名称": ("name", None, True),
            "桥梁类别": ("category_code", "category", True),
            "评定单元": ("assessment_unit_code", "assessment_unit", False),
            "桥梁类型": ("bridge_type_code", "bridge_type", True),
            "部位": ("part_code", "part", True),
            "结构类型": ("structure_code", "structure", False),
            "部件类型": ("component_type_code", "component_type", False),
            "构件形式": ("component_form_code", "component_form", False),
            "病害类型": ("disease_code", "disease", True),
            "标度": ("scale_code", "scale", True),
            "定性描述": ("quality_code", "quality", False),
            "定量描述": ("quantity_code", "quantity", False),
        }

        # 验证列
        for col_name, (field_name, ref_key, is_required) in column_mapping.items():
            # 获取行数据的值
            value = (
                str(row.get(col_name, "")).strip()
                if pd.notna(row.get(col_name))
                else ""
            )

            if field_name == "code":
                result["data"][field_name] = value
            elif field_name == "name":
                if not value:
                    result["errors"].append(
                        {
                            "column": col_name,
                            "type": "required",
                            "message": "名称不能为空",
                        }
                    )
                    result["is_valid"] = False
                else:
                    result["data"][field_name] = value
            else:
                # 其他字段匹配参考数据
                if value:
                    match_result = match_name_to_code_func(
                        value, ref_key, reference_data
                    )
                    if match_result["matched"]:
                        result["data"][field_name] = match_result["code"]
                    else:
                        result["errors"].append(
                            {
                                "column": col_name,
                                "type": "not_found",
                                "message": f"找不到匹配的{col_name}: '{value}'",
                            }
                        )
                        result["is_valid"] = False
                elif is_required:
                    result["errors"].append(
                        {
                            "column": col_name,
                            "type": "required",
                            "message": f"{col_name}不能为空",
                        }
                    )
                    result["is_valid"] = False
                else:
                    result["data"][field_name] = None

        return result

    except Exception as e:
        result["is_valid"] = False
        result["errors"].append(
            {
                "column": "全行",
                "type": "validation_error",
                "message": f"验证时出错: {str(e)}",
            }
        )
        return result
