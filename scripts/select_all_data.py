import pandas as pd
import os
from openpyxl import load_workbook


def read_excel_file(file_path):
    """
    读取Excel文件并遍历所有工作表
    """
    if not os.path.exists(file_path):
        print(f"文件 {file_path} 不存在!")
        return

    try:
        # 方法1: 使用pandas读取所有工作表
        print("=== 使用 pandas 读取 ===")
        excel_data = pd.read_excel(file_path, sheet_name=None)

        print(f"文件: {file_path}")
        print(f"工作表数量: {len(excel_data)}")
        print(f"工作表名称: {list(excel_data.keys())}")
        print("-" * 50)

        # 遍历每个工作表
        for sheet_name, df in excel_data.items():
            print(f"\n工作表: {sheet_name}")
            print(f"行数: {len(df)}")
            print(f"列数: {len(df.columns)}")
            print(f"列名: {list(df.columns)}")

            # 显示前几行数据
            print("前5行数据:")
            print(df.head())
            print("-" * 30)

        # 方法2: 使用openpyxl获取更详细信息
        print("\n=== 使用 openpyxl 获取详细信息 ===")
        workbook = load_workbook(file_path, read_only=True)

        print(f"所有工作表: {workbook.sheetnames}")

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"\n工作表: {sheet_name}")
            print(f"最大行数: {sheet.max_row}")
            print(f"最大列数: {sheet.max_column}")

            # 读取表头
            if sheet.max_row > 0:
                headers = []
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=1, column=col).value
                    headers.append(cell_value)
                print(f"表头: {headers}")

        workbook.close()

    except Exception as e:
        print(f"读取文件时出错: {e}")


def analyze_specific_sheet(file_path, sheet_name):
    """
    分析特定工作表的详细信息
    """
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        print(f"\n=== {sheet_name} 详细分析 ===")
        print(f"数据形状: {df.shape}")
        print(f"数据类型:")
        print(df.dtypes)
        print(f"\n数据概览:")
        print(df.describe())
        print(f"\n缺失值:")
        print(df.isnull().sum())

        return df

    except Exception as e:
        print(f"分析工作表 {sheet_name} 时出错: {e}")
        return None


def get_available_sheets(file_path):
    """
    获取所有可用的工作表
    """
    try:
        excel_data = pd.read_excel(file_path, sheet_name=None)
        return list(excel_data.keys())
    except Exception as e:
        print(f"读取文件时出错: {e}")
        return []


def extract_hierarchical_structure(df, sheet_name):
    """
    提取深层次的层级关系结构 - 改进版
    """
    if df is None:
        return None

    print(f"\n=== 提取 {sheet_name} 结构层级关系 ===")

    # 获取列名
    columns = df.columns.tolist()
    print(f"表格列数: {len(columns)}")
    print(f"原始列名: {columns}")

    # 跳过第一行（标题行），从第二行开始作为列标题
    if len(df) > 1:
        # 使用第二行作为列标题
        header_row = df.iloc[1].fillna("")
        print(f"第二行内容（列标题）: {header_row.tolist()}")

        # 从第三行开始是数据
        data_df = df.iloc[2:].copy()
        data_df.columns = header_row
    else:
        print("数据行数不足，使用原始列名")
        data_df = df.copy()

    # 清理列名
    clean_columns = []
    for col in data_df.columns:
        if pd.isna(col) or col == "":
            clean_columns.append(f"未命名_{len(clean_columns)}")
        else:
            clean_columns.append(str(col))

    data_df.columns = clean_columns
    print(f"清理后的列名: {clean_columns}")

    # 定义标准的列名映射
    expected_columns = [
        "桥梁类型",
        "部位",
        "结构类型",
        "部件类型",
        "构件形式",
        "病害类型",
        "标度",
        "定性描述",
        "定量描述",
    ]

    # 确保我们有足够的列
    if len(data_df.columns) >= len(expected_columns):
        data_df.columns = expected_columns + list(
            data_df.columns[len(expected_columns) :]
        )
    else:
        # 如果列数不够，用已有的列
        data_df.columns = expected_columns[: len(data_df.columns)]

    print(f"最终列名: {list(data_df.columns)}")

    # 重置索引
    data_df = data_df.reset_index(drop=True)

    # 处理合并单元格 - 向下填充空值
    hierarchy_columns = [
        "桥梁类型",
        "部位",
        "结构类型",
        "部件类型",
        "构件形式",
        "病害类型",
    ]

    print("\n=== 处理合并单元格，向下填充空值 ===")
    for col in hierarchy_columns:
        if col in data_df.columns:
            # 清理数据：将空字符串和None转换为NaN
            data_df[col] = data_df[col].replace("", pd.NA)
            data_df[col] = data_df[col].where(data_df[col].notna(), pd.NA)

            # 向下填充
            data_df[col] = data_df[col].fillna(method="ffill")

            print(f"{col} 列处理完成，非空值数量: {data_df[col].notna().sum()}")

    # 显示处理后的前几行数据以便调试
    print("\n=== 处理后的前10行数据 ===")
    for i, row in data_df.head(10).iterrows():
        print(f"行{i}: {dict(row[hierarchy_columns])}")

    # 构建层级结构
    hierarchy = build_enhanced_nested_structure(data_df, hierarchy_columns)

    return hierarchy


def build_enhanced_nested_structure(data_df, hierarchy_columns):
    """
    构建增强的嵌套层级结构
    """
    print(f"\n=== 构建嵌套结构 ===")
    result = {}

    # 详细信息列
    detail_columns = ["标度", "定性描述", "定量描述"]

    # 按桥梁类型分组
    bridge_types = data_df["桥梁类型"].dropna().unique()
    bridge_types = [bt for bt in bridge_types if str(bt).strip() and str(bt) != "nan"]

    print(f"发现的桥梁类型: {bridge_types}")

    for bridge_type in bridge_types:
        print(f"\n处理桥梁类型: {bridge_type}")

        # 筛选属于该桥梁类型的数据
        bridge_data = data_df[data_df["桥梁类型"] == bridge_type].copy()
        print(f"该桥梁类型的数据行数: {len(bridge_data)}")

        # 初始化桥梁类型的结构
        result[bridge_type] = {}

        # 构建多层级结构，现在包括病害类型
        build_recursive_structure(
            bridge_data,
            result[bridge_type],
            hierarchy_columns[
                1:
            ],  # 跳过桥梁类型，包含：部位、结构类型、部件类型、构件形式、病害类型
            detail_columns,
            0,
        )

    return result


def build_recursive_structure(
    data, current_dict, remaining_columns, detail_columns, depth
):
    """
    递归构建层级结构 - 修复版
    """
    if not remaining_columns:
        # 已经到达最深层（病害类型），需要构建详细信息的层级结构
        print(f"{'  ' * depth}到达最深层，构建详细信息层级")
        build_detail_structure(data, current_dict, detail_columns, depth)
        return

    current_column = remaining_columns[0]
    remaining = remaining_columns[1:]

    # 获取当前层级的所有唯一值
    unique_values = data[current_column].dropna().unique()
    unique_values = [
        val for val in unique_values if str(val).strip() and str(val) != "nan"
    ]

    print(
        f"{'  ' * depth}第{depth+2}层 ({current_column}): {len(unique_values)}个选项 - {unique_values}"
    )

    for value in unique_values:
        # 筛选属于该值的数据
        filtered_data = data[data[current_column] == value]

        if len(filtered_data) > 0:
            current_dict[value] = {}

            # 递归处理下一层
            build_recursive_structure(
                filtered_data, current_dict[value], remaining, detail_columns, depth + 1
            )


def build_detail_structure(data, current_dict, detail_columns, depth):
    """
    构建详细信息的层级结构：标度 → 定性描述 → 定量描述
    """
    print(f"{'  ' * depth}构建详细信息层级，数据行数: {len(data)}")

    # 先按标度分组
    scale_values = data["标度"].dropna().unique()
    scale_values = [
        val for val in scale_values if str(val).strip() and str(val) != "nan"
    ]

    print(f"{'  ' * depth}发现标度值: {scale_values}")

    for scale in scale_values:
        scale_data = data[data["标度"] == scale]
        current_dict[str(scale)] = {}

        # 按定性描述分组
        qual_values = scale_data["定性描述"].dropna().unique()
        qual_values = [
            val for val in qual_values if str(val).strip() and str(val) != "nan"
        ]

        for qual_desc in qual_values:
            qual_data = scale_data[scale_data["定性描述"] == qual_desc]
            current_dict[str(scale)][str(qual_desc)] = {}

            # 按定量描述分组
            quan_values = qual_data["定量描述"].dropna().unique()
            quan_values = [
                val for val in quan_values if str(val).strip() and str(val) != "nan"
            ]

            for quan_desc in quan_values:
                current_dict[str(scale)][str(qual_desc)][str(quan_desc)] = "完整信息"


def interactive_multilevel_analysis(hierarchy, sheet_name):
    """
    多层级交互式分析 - 修复版
    """
    if not hierarchy:
        print("没有可分析的数据")
        return

    # 完整的9层结构
    levels = [
        "部位",
        "结构类型",
        "部件类型",
        "构件形式",
        "病害类型",
        "标度",
        "定性描述",
        "定量描述",
    ]

    while True:
        print("\n" + "=" * 60)
        print(f"=== {sheet_name} 多层级分析系统 ===")
        print("=" * 60)

        # 第一层：选择桥梁类型
        bridge_types = list(hierarchy.keys())
        if not bridge_types:
            print("没有找到桥梁类型数据")
            break

        print("\n【第一层】请选择桥梁类型:")
        for i, bridge_type in enumerate(bridge_types, 1):
            print(f"  {i}. {bridge_type}")
        print(f"  0. 返回工作表选择")

        try:
            choice = input("\n请输入选择 (数字): ").strip()
            if choice == "0":
                break

            bridge_idx = int(choice) - 1
            if 0 <= bridge_idx < len(bridge_types):
                selected_bridge = bridge_types[bridge_idx]
                print(f"\n✓ 已选择桥梁类型: {selected_bridge}")

                # 开始多层级选择
                current_data = hierarchy[selected_bridge]
                current_path = [selected_bridge]

                if navigate_enhanced_levels(current_data, current_path, levels, 0):
                    continue  # 如果返回True，回到桥梁类型选择
                else:
                    break  # 如果返回False，退出到工作表选择
            else:
                print("无效选择，请重新输入")

        except ValueError:
            print("请输入有效的数字")
        except KeyboardInterrupt:
            print("\n程序被用户中断")
            break


def navigate_enhanced_levels(data, path, levels, level_index):
    """
    增强的递归导航各个层级 - 修复版
    """
    print(f"\n=== 调试信息 ===")
    print(f"当前层级索引: {level_index}")
    print(f"当前路径: {' → '.join(path)}")
    print(
        f"当前数据键: {list(data.keys()) if isinstance(data, dict) else '非字典类型'}"
    )
    print(f"剩余层级: {levels[level_index:] if level_index < len(levels) else '无'}")

    if level_index >= len(levels):
        # 已经完成所有选择，显示最终确认
        print(f"\n🎉 完整路径选择完成！")
        print(f"完整路径: {' → '.join(path)}")
        print("-" * 60)
        input("\n选择完成！按回车键返回工作表选择...")
        return False  # 返回False表示直接退出到工作表选择

    current_level_name = levels[level_index]

    while True:
        print(f"\n【第{level_index + 2}层】请选择{current_level_name}:")

        # 检查数据类型
        if not isinstance(data, dict):
            print(f"数据类型错误，期望字典，实际: {type(data)}")
            return True

        # 获取当前层级的选项
        options = list(data.keys())

        if not options:
            print(f"没有可选的{current_level_name}")
            print(f"当前数据包含: {list(data.keys())}")
            input("\n按回车键返回上一层...")
            return True

        # 显示选项
        for i, option in enumerate(options, 1):
            # 根据层级显示不同的统计信息
            sub_data = data[option]
            if isinstance(sub_data, dict):
                if level_index == len(levels) - 1:  # 最后一层（定量描述）
                    print(f"  {i}. {option}")
                else:
                    # 统计下一层的选项数量
                    next_level = (
                        levels[level_index + 1]
                        if level_index + 1 < len(levels)
                        else "项目"
                    )
                    sub_count = len(sub_data) if isinstance(sub_data, dict) else 0
                    print(f"  {i}. {option} ({sub_count}个{next_level})")
            else:
                print(f"  {i}. {option}")

        print(f"  0. 返回上一层")

        try:
            choice = input(f"\n请输入选择 (数字): ").strip()
            if choice == "0":
                return True

            option_idx = int(choice) - 1
            if 0 <= option_idx < len(options):
                selected_option = options[option_idx]
                new_path = path + [selected_option]
                print(f"\n✓ 已选择{current_level_name}: {selected_option}")
                print(f"当前路径: {' → '.join(new_path)}")

                # 检查是否已经选择完所有层级
                if level_index == len(levels) - 1:
                    # 已经是最后一层，完成选择
                    print(f"\n🎉 完整路径选择完成！")
                    print(f"最终路径: {' → '.join(new_path)}")
                    print("-" * 60)
                    input("\n选择完成！按回车键返回工作表选择...")
                    return False  # 直接返回到工作表选择
                else:
                    # 递归到下一层
                    result = navigate_enhanced_levels(
                        data[selected_option],
                        new_path,
                        levels,
                        level_index + 1,
                    )
                    if not result:  # 如果下层返回False，向上传递
                        return False
                    # 如果下层返回True，继续当前层的循环
            else:
                print("无效选择，请重新输入")

        except ValueError:
            print("请输入有效的数字")
        except KeyboardInterrupt:
            print("\n程序被用户中断")
            return False


# 删除了 display_enhanced_details 函数，因为新逻辑不再需要它


def interactive_bridge_analysis(file_path):
    """
    交互式桥梁分析主函数
    """
    sheet_names = get_available_sheets(file_path)

    if not sheet_names:
        print("无法读取工作表")
        return

    while True:
        print("\n" + "=" * 60)
        print("=== 桥梁数据分析系统 ===")
        print("=" * 60)
        print("请选择要分析的工作表:")
        for i, sheet_name in enumerate(sheet_names, 1):
            print(f"  {i}. {sheet_name}")
        print("  0. 退出程序")

        choice = input("\n请输入选择 (数字): ").strip()

        if choice == "0":
            print("程序结束，再见！")
            break
        elif choice.isdigit() and 1 <= int(choice) <= len(sheet_names):
            selected_sheet = sheet_names[int(choice) - 1]
            print(f"\n=== 正在分析工作表: {selected_sheet} ===")

            try:
                # 读取选定的工作表
                df = pd.read_excel(file_path, sheet_name=selected_sheet, header=None)
                print(f"数据形状: {df.shape}")
                print("正在提取层级关系...")

                # 提取层级关系
                hierarchy = extract_hierarchical_structure(df, selected_sheet)

                if hierarchy:
                    print("\n=== 数据提取完成 ===")
                    print(f"共发现 {len(hierarchy)} 种桥梁类型:")
                    for bridge_type in hierarchy.keys():
                        print(f"  • {bridge_type}")

                    # 启动交互式多层级分析
                    interactive_multilevel_analysis(hierarchy, selected_sheet)
                else:
                    print("数据提取失败，可能是表格结构不符合预期")
            except Exception as e:
                print(f"处理工作表时出错: {e}")
                import traceback

                traceback.print_exc()
        else:
            print("无效选择，请重新输入")


def main():
    # Excel文件路径
    excel_file = "work.xls"

    # 检查文件是否存在
    if os.path.exists(excel_file):
        print("找到 work.xls 文件!")

        # 交互式分析
        interactive_bridge_analysis(excel_file)

    else:
        print("work.xls 文件不存在，请确保文件在当前目录下")

        # 显示当前目录下的所有文件
        print("\n当前目录下的文件:")
        for file in os.listdir("."):
            print(f"  {file}")


if __name__ == "__main__":
    # 安装依赖提示
    print("请确保已安装必要的库:")
    print("pip install pandas openpyxl xlrd")
    print("-" * 50)

    main()
